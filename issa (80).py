# ==========================================
# 0. CONNEXION SUPABASE (POSTGRESQL)
# ==========================================

import base64
from datetime import datetime
import io
import json
import os
import zipfile
import unicodedata
import numpy as np
import pandas as pd
from fpdf import FPDF
import streamlit as st
import bcrypt
from supabase import create_client, Client

# --- Configuration Supabase ---
SUPABASE_URL = "https://dzxotavktglasrcpyrwx.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImR6eG90YXZrdGdsYXNyY3B5cnd4Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODY3MjAyODEsImV4cCI6MjEwMjI5NjI4MX0.gYODj-fzOSs6BiWewTMln-NPfaESN5EABVH0xmJrtLw"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# --- Tables Supabase ---
TABLE_ADMINS = "admins"
TABLE_PROFESSEURS = "professeurs"
TABLE_ELEVES = "eleves"
TABLE_CLASSES = "classes"
TABLE_MATIERES = "matieres"
TABLE_NOTES = "notes"
TABLE_PERIODES = "periodes"
TABLE_VIE_SCOLAIRE = "vie_scolaire"
TABLE_TRAVAILS = "travaux"
TABLE_MESSAGES = "messages"
TABLE_AUDIT_LOGS = "audit_logs"

# --- Sécurité ---
def hacher_mot_de_passe(password: str) -> str:
    if not password:
        return ""
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")


def verifier_mot_de_passe(password: str, hashed: str) -> bool:
    if not password or not hashed:
        return False
    try:
        return bcrypt.checkpw(password.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def normaliser_texte(texte):
    if not texte:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", str(texte))
        if unicodedata.category(c) != "Mn"
    ).strip().lower()


def nettoyer_texte_pdf(texte):
    if not texte:
        return ""
    return str(texte).encode("latin-1", "replace").decode("latin-1")


ADMIN_EMAIL = "cpnm@gmail.com"

# --- Journalisation Supabase ---
def enregistrer_log_action(acteur: str, action: str, details: str):
    try:
        supabase.table(TABLE_AUDIT_LOGS).insert({
            "horodatage": datetime.now().isoformat(),
            "acteur": acteur,
            "action": action,
            "details": details
        }).execute()
    except Exception:
        pass


def trier_eleves_par_nom(df):
    if df is None or df.empty:
        return df
    df_copy = df.copy()
    if "Nom" in df_copy.columns and "Prénom" in df_copy.columns:
        df_copy["Nom_Sort"] = df_copy["Nom"].astype(str).str.strip().str.upper()
        df_copy["Prenom_Sort"] = df_copy["Prénom"].astype(str).str.strip().str.upper()
        df_copy = (
            df_copy.sort_values(by=["Nom_Sort", "Prenom_Sort"])
            .drop(columns=["Nom_Sort", "Prenom_Sort"])
        )
    return df_copy.reset_index(drop=True)


def synchroniser_listes_blanches():
    pass# ==========================================
# 2. INITIALISATION DES DONNÉES (SUPABASE)
# ==========================================

# --- Fonctions génériques Supabase ---
def charger_table(table_name):
    try:
        response = supabase.table(table_name).select("*").execute()
        data = response.data if response.data else []
        return pd.DataFrame(data)
    except Exception:
        return pd.DataFrame()


def sauvegarder_ligne(table_name, data):
    try:
        supabase.table(table_name).insert(data).execute()
        return True
    except Exception:
        return False


def mettre_a_jour_ligne(table_name, filtre, data):
    try:
        query = supabase.table(table_name).update(data)
        for cle, valeur in filtre.items():
            query = query.eq(cle, valeur)
        query.execute()
        return True
    except Exception:
        return False


def supprimer_ligne(table_name, filtre):
    try:
        query = supabase.table(table_name).delete()
        for cle, valeur in filtre.items():
            query = query.eq(cle, valeur)
        query.execute()
        return True
    except Exception:
        return False


# --- Chargement des tables ---
admins_db = charger_table(TABLE_ADMINS)
professeurs_db = charger_table(TABLE_PROFESSEURS)
eleves_db = charger_table(TABLE_ELEVES)
classes_db = charger_table(TABLE_CLASSES)
matieres_db = charger_table(TABLE_MATIERES)
notes_db = charger_table(TABLE_NOTES)
periodes_db = charger_table(TABLE_PERIODES)
vie_scolaire_db = charger_table(TABLE_VIE_SCOLAIRE)
travaux_db = charger_table(TABLE_TRAVAILS)
messages_db = charger_table(TABLE_MESSAGES)
audit_logs_db = charger_table(TABLE_AUDIT_LOGS)

# --- Initialisation des états Streamlit ---
if "espace_actif" not in st.session_state:
    st.session_state.espace_actif = "🏠 Accueil"

if "authenticated_admin" not in st.session_state:
    st.session_state.authenticated_admin = False

if "edt_documents" not in st.session_state:
    st.session_state.edt_documents = {}

# --- Synchronisation avec Supabase ---
st.session_state.admin_credentials = admins_db
st.session_state.prof_credentials = professeurs_db
st.session_state.eleves_db = eleves_db
st.session_state.classes_db = classes_db
st.session_state.matieres_def = matieres_db
st.session_state.coefficients_db = matieres_db.copy()
st.session_state.notes_db = notes_db
st.session_state.periodes_db = periodes_db
st.session_state.viescolaire_db = vie_scolaire_db
st.session_state.travail_a_faire_db = travaux_db
st.session_state.messages_parents_db = messages_db
st.session_state.audit_logs_db = audit_logs_db

# --- Création de l’administrateur si absent ---
if admins_db.empty:
    admin = {
        "nom": "Principal",
        "prenom": "Admin",
        "email": ADMIN_EMAIL,
        "mot_de_passe": hacher_mot_de_passe("cpnm2026"),
        "niveau_acces": "Super-Admin Ayant-Droit"
    }
    sauvegarder_ligne(TABLE_ADMINS, admin)
    st.session_state.admin_credentials = charger_table(TABLE_ADMINS)

# --- Tri des élèves ---
if not st.session_state.eleves_db.empty:
    st.session_state.eleves_db = trier_eleves_par_nom(
        st.session_state.eleves_db
    )

# --- Listes blanches ---
st.session_state.admin_white_list = st.session_state.admin_credentials.copy()
st.session_state.prof_white_list = st.session_state.prof_credentials.copy()

# --- Emploi du temps ---
JOURS_LIST = [
    "Lundi",
    "Mardi",
    "Mercredi",
    "Jeudi",
    "Vendredi",
    "Samedi",
]

HEURES_LIST = [
    "08h-09h",
    "09h-10h",
    "10h-11h",
    "11h00-11h30",
    "11h30-12h",
    "12h-13h",
    "13h-14h",
    "14h-15h",
    "15h-16h",
    "16h-17h",
    "17h-18h",
    "18h-19h",
]

if "edt_grid_db" not in st.session_state:
    st.session_state.edt_grid_db = {}


def get_or_create_edt(classe):
    if classe not in st.session_state.edt_grid_db:
        df = pd.DataFrame("", index=JOURS_LIST, columns=HEURES_LIST)
        df["11h00-11h30"] = "Récréation"
        st.session_state.edt_grid_db[classe] = df
    return st.session_state.edt_grid_db[classe]


# --- Cahier de textes ---
if "cahier_textes" not in st.session_state:
    st.session_state.cahier_textes = pd.DataFrame(
        columns=[
            "Professeur",
            "Date",
            "Classe",
            "Matière",
            "Contenu",
            "Travail à faire",
        ]
    )

# --- Absences ---
if "absences_db" not in st.session_state:
    st.session_state.absences_db = pd.DataFrame(
        columns=["Date", "Classe", "Élève", "Statut", "Motif"]
    )

synchroniser_listes_blanches()# ==========================================
# 3. FONCTIONS MÉTIER (SUPABASE)
# ==========================================

# ---------- CLASSES ----------

def obtenir_cycle_classe(classe_nom):
    if not classe_nom:
        return "Élémentaire"

    try:
        res = (
            supabase.table(TABLE_CLASSES)
            .select("cycle")
            .eq("classe", classe_nom)
            .limit(1)
            .execute()
        )

        if res.data:
            return res.data[0]["cycle"]
    except Exception:
        pass

    classe = str(classe_nom).upper()

    if any(x in classe for x in ["6ÈME", "6EME", "5ÈME", "5EME", "4ÈME", "4EME", "3ÈME", "3EME"]):
        return "Collège"

    return "Élémentaire"


def est_cycle_elementaire(cycle):
    if not cycle:
        return True

    return "élément" in str(cycle).lower() or "element" in str(cycle).lower()


# ---------- PÉRIODES ----------

def obtenir_periodes_pour_classe(classe_nom):
    cycle = obtenir_cycle_classe(classe_nom)

    try:
        res = (
            supabase.table(TABLE_PERIODES)
            .select("*")
            .eq("cycle", cycle)
            .execute()
        )

        if res.data:
            return [x["periode"] for x in res.data]
    except Exception:
        pass

    if est_cycle_elementaire(cycle):
        return ["1er Trimestre", "2ème Trimestre", "3ème Trimestre"]

    return ["1er Semestre", "2ème Semestre"]


# ---------- MATIÈRES ----------

def obtenir_coefficient_matiere(classe, matiere):
    try:
        res = (
            supabase.table(TABLE_MATIERES)
            .select("coefficient")
            .eq("classe", classe)
            .eq("matiere", matiere)
            .limit(1)
            .execute()
        )

        if res.data:
            return float(res.data[0]["coefficient"])
    except Exception:
        pass

    return 1.0


def obtenir_bareme_matiere(classe, matiere):
    try:
        res = (
            supabase.table(TABLE_MATIERES)
            .select("bareme")
            .eq("classe", classe)
            .eq("matiere", matiere)
            .limit(1)
            .execute()
        )

        if res.data:
            return float(res.data[0]["bareme"])
    except Exception:
        pass

    return 20.0


# ---------- APPRÉCIATION ----------

def obtenir_appreciation(moyenne):
    if moyenne >= 18:
        return "Excellent"
    elif moyenne >= 16:
        return "Très Bien"
    elif moyenne >= 14:
        return "Bien"
    elif moyenne >= 12:
        return "Assez Bien"
    elif moyenne >= 10:
        return "Passable"
    elif moyenne >= 8:
        return "Insuffisant"
    else:
        return "Faible"


# ---------- ÉLÈVES ----------

def charger_eleves(classe=None):
    try:
        query = supabase.table(TABLE_ELEVES).select("*")

        if classe:
            query = query.eq("classe", classe)

        res = query.execute()

        return pd.DataFrame(res.data)

    except Exception:
        return pd.DataFrame()


def ajouter_eleve(data):
    return sauvegarder_ligne(TABLE_ELEVES, data)


def modifier_eleve(id_eleve, data):
    return mettre_a_jour_ligne(TABLE_ELEVES, {"id": id_eleve}, data)


def supprimer_eleve(id_eleve):
    return supprimer_ligne(TABLE_ELEVES, {"id": id_eleve})


# ---------- NOTES ----------

def charger_notes(classe=None, periode=None, eleve=None):
    try:
        query = supabase.table(TABLE_NOTES).select("*")

        if classe:
            query = query.eq("classe", classe)

        if periode:
            query = query.eq("periode", periode)

        if eleve:
            query = query.eq("eleve", eleve)

        res = query.execute()

        return pd.DataFrame(res.data)

    except Exception:
        return pd.DataFrame()


def enregistrer_note(classe, periode, eleve, matiere, devoir1, devoir2, composition):
    try:
        exist = (
            supabase.table(TABLE_NOTES)
            .select("id")
            .eq("classe", classe)
            .eq("periode", periode)
            .eq("eleve", eleve)
            .eq("matiere", matiere)
            .limit(1)
            .execute()
        )

        data = {
            "classe": classe,
            "periode": periode,
            "eleve": eleve,
            "matiere": matiere,
            "devoir1": devoir1,
            "devoir2": devoir2,
            "composition": composition,
        }

        if exist.data:
            id_note = exist.data[0]["id"]

            supabase.table(TABLE_NOTES).update(data).eq("id", id_note).execute()
        else:
            supabase.table(TABLE_NOTES).insert(data).execute()

        return True

    except Exception:
        return False


# ---------- CALCUL MOYENNE ----------

def calculer_moyenne_eleve(classe, eleve, periode):
    notes = charger_notes(classe, periode, eleve)

    if notes.empty:
        return 0.0

    total = 0
    coef_total = 0

    for _, n in notes.iterrows():
        coef = obtenir_coefficient_matiere(classe, n["matiere"])

        d1 = float(n.get("devoir1", 0) or 0)
        d2 = float(n.get("devoir2", 0) or 0)
        comp = float(n.get("composition", 0) or 0)

        moyenne = ((d1 + d2) / 2 + comp) / 2

        total += moyenne * coef
        coef_total += coef

    if coef_total == 0:
        return 0.0

    return round(total / coef_total, 2)


# ---------- CLASSEMENT ----------

def classement_classe(classe, periode):
    eleves = charger_eleves(classe)

    if eleves.empty:
        return []

    classement = []

    for _, e in eleves.iterrows():
        moyenne = calculer_moyenne_eleve(classe, e["nom_complet"], periode)

        classement.append({
            "eleve": e["nom_complet"],
            "moyenne": moyenne,
        })

    classement.sort(key=lambda x: x["moyenne"], reverse=True)

    return classement# ==========================================
# 4. BULLETINS (SUPABASE)
# ==========================================

def calculer_bulletin_eleve(classe, eleve, periode):
    cycle = obtenir_cycle_classe(classe)
    is_elem = est_cycle_elementaire(cycle)

    notes = charger_notes(classe, periode, eleve)

    if notes.empty:
        return None

    lignes = []
    total_points = 0.0
    total_coeff = 0.0
    total_bareme = 0.0

    for _, note in notes.iterrows():
        matiere = note["matiere"]
        coef = obtenir_coefficient_matiere(classe, matiere)
        bareme = obtenir_bareme_matiere(classe, matiere)

        d1 = float(note.get("devoir1", 0) or 0)
        d2 = float(note.get("devoir2", 0) or 0)
        comp = float(note.get("composition", 0) or 0)

        if is_elem:
            moyenne = comp
            total_points += moyenne
            total_bareme += bareme

            lignes.append({
                "Matiere": matiere,
                "Bareme": bareme,
                "Composition": comp,
                "MoyenneMatiere": round(moyenne, 2),
                "Appreciation": obtenir_appreciation(
                    (moyenne / bareme) * 20 if bareme else moyenne
                ),
            })
        else:
            moyenne = ((d1 + d2) / 2 + comp) / 2
            total = moyenne * coef

            total_points += total
            total_coeff += coef

            lignes.append({
                "Matiere": matiere,
                "Coefficient": coef,
                "Devoir1": d1,
                "Devoir2": d2,
                "Composition": comp,
                "MoyenneMatiere": round(moyenne, 2),
                "TotalPondere": round(total, 2),
                "Appreciation": obtenir_appreciation(moyenne),
            })

    if is_elem:
        moyenne_generale = (
            round((total_points / total_bareme) * 10, 2)
            if total_bareme > 0 else 0
        )
    else:
        moyenne_generale = (
            round(total_points / total_coeff, 2)
            if total_coeff > 0 else 0
        )

    classement = classement_classe(classe, periode)

    rang = "-"
    effectif = len(classement)

    for i, e in enumerate(classement, 1):
        if e["eleve"] == eleve:
            rang = f"{i} / {effectif}"
            break

    abs_just = 0
    abs_non_just = 0
    retards = 0
    heures_perdues = 0
    observations = "RAS"
    decision = "Encouragements"

    try:
        vs = (
            supabase.table(TABLE_VIE_SCOLAIRE)
            .select("*")
            .eq("classe", classe)
            .eq("periode", periode)
            .eq("eleve", eleve)
            .limit(1)
            .execute()
        )

        if vs.data:
            v = vs.data[0]
            abs_just = int(v.get("absences_justifiees", 0))
            abs_non_just = int(v.get("absences_non_justifiees", 0))
            retards = int(v.get("retards", 0))
            heures_perdues = int(v.get("heures_perdues", 0))
            observations = v.get("observations", "RAS")
            decision = v.get("decision_conseil", "Encouragements")
    except Exception:
        pass

    return {
        "eleve": eleve,
        "classe": classe,
        "cycle": cycle,
        "periode": periode,
        "lignes": lignes,
        "total_points": round(total_points, 2),
        "total_coefficients": total_coeff if not is_elem else "-",
        "total_bareme": 10.0 if is_elem else 20.0,
        "moyenne_generale": moyenne_generale,
        "rang": rang,
        "effectif": effectif,
        "abs_just": abs_just,
        "abs_non_just": abs_non_just,
        "retards": retards,
        "heures_perdues": heures_perdues,
        "observations": observations,
        "decision": decision,
    }


def generer_pdf_bulletin(bul_data):
    pdf = FPDF()

    try:
        if os.path.exists("DejaVuSans.ttf"):
            pdf.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
            pdf.add_font("DejaVu", "B", "DejaVuSans-Bold.ttf", uni=True)
            font = "DejaVu"
        else:
            font = "Arial"
    except Exception:
        font = "Arial"

    pdf.add_page()

    ajouter_entete_senegal_officiel(
        pdf,
        f"BULLETIN DE NOTES - {bul_data['periode']}"
    )

    pdf.set_font(font, "B", 11)
    pdf.cell(0, 7, nettoyer_texte_pdf(
        f"Élève : {bul_data['eleve']}"
    ), 0, 1)

    pdf.cell(0, 7, nettoyer_texte_pdf(
        f"Classe : {bul_data['classe']}"
    ), 0, 1)

    pdf.cell(0, 7, nettoyer_texte_pdf(
        f"Rang : {bul_data['rang']} / {bul_data['effectif']}"
    ), 0, 1)

    pdf.ln(4)

    cycle = bul_data["cycle"]

    if est_cycle_elementaire(cycle):
        pdf.set_font(font, "B", 9)
        pdf.cell(90, 7, "Matière", 1)
        pdf.cell(25, 7, "Barème", 1)
        pdf.cell(30, 7, "Note", 1)
        pdf.cell(45, 7, "Appréciation", 1)
        pdf.ln()

        pdf.set_font(font, "", 9)

        for l in bul_data["lignes"]:
            pdf.cell(90, 7, nettoyer_texte_pdf(l["Matiere"]), 1)
            pdf.cell(25, 7, str(l["Bareme"]), 1, 0, "C")
            pdf.cell(30, 7, str(l["Composition"]), 1, 0, "C")
            pdf.cell(45, 7, nettoyer_texte_pdf(l["Appreciation"]), 1)
            pdf.ln()
    else:
        pdf.set_font(font, "B", 8)
        pdf.cell(60, 7, "Matière", 1)
        pdf.cell(15, 7, "Coef", 1)
        pdf.cell(20, 7, "D1", 1)
        pdf.cell(20, 7, "D2", 1)
        pdf.cell(25, 7, "Comp", 1)
        pdf.cell(25, 7, "Moy", 1)
        pdf.ln()

        pdf.set_font(font, "", 8)

        for l in bul_data["lignes"]:
            pdf.cell(60, 7, nettoyer_texte_pdf(l["Matiere"]), 1)
            pdf.cell(15, 7, str(l["Coefficient"]), 1, 0, "C")
            pdf.cell(20, 7, str(l["Devoir1"]), 1, 0, "C")
            pdf.cell(20, 7, str(l["Devoir2"]), 1, 0, "C")
            pdf.cell(25, 7, str(l["Composition"]), 1, 0, "C")
            pdf.cell(25, 7, str(l["MoyenneMatiere"]), 1, 0, "C")
            pdf.ln()

    pdf.ln(4)

    pdf.set_font(font, "B", 11)
    pdf.cell(0, 8, nettoyer_texte_pdf(
        f"Moyenne Générale : {bul_data['moyenne_generale']}"
    ), 0, 1)

    pdf.set_font(font, "", 9)
    pdf.cell(0, 6, nettoyer_texte_pdf(
        f"Absences justifiées : {bul_data['abs_just']}"
    ), 0, 1)

    pdf.cell(0, 6, nettoyer_texte_pdf(
        f"Absences non justifiées : {bul_data['abs_non_just']}"
    ), 0, 1)

    pdf.cell(0, 6, nettoyer_texte_pdf(
        f"Retards : {bul_data['retards']}"
    ), 0, 1)

    pdf.multi_cell(
        0,
        6,
        nettoyer_texte_pdf(
            f"Observations : {bul_data['observations']}"
        ),
    )

    pdf.cell(0, 6, nettoyer_texte_pdf(
        f"Décision : {bul_data['decision']}"
    ), 0, 1)

    ajouter_bloc_signatures(pdf)

    return pdf.output(dest="S").encode("latin-1")# ==========================================
# 5. AUTHENTIFICATION (SUPABASE)
# ==========================================

# ---------- ADMIN ----------

def connexion_admin(email, mot_de_passe):
    try:
        res = (
            supabase.table(TABLE_ADMINS)
            .select("*")
            .eq("email", email)
            .limit(1)
            .execute()
        )

        if not res.data:
            return False

        admin = res.data[0]

        if verifier_mot_de_passe(mot_de_passe, admin["mot_de_passe"]):
            st.session_state.authenticated_admin = True
            st.session_state.admin = admin

            enregistrer_log_action(
                email,
                "Connexion Admin",
                "Connexion réussie"
            )

            return True

    except Exception:
        pass

    return False


def deconnexion_admin():
    st.session_state.authenticated_admin = False
    st.session_state.admin = None


# ---------- PROFESSEUR ----------

def connexion_professeur(email, mot_de_passe):
    try:
        res = (
            supabase.table(TABLE_PROFESSEURS)
            .select("*")
            .eq("email", email)
            .limit(1)
            .execute()
        )

        if not res.data:
            return False

        prof = res.data[0]

        if verifier_mot_de_passe(mot_de_passe, prof["mot_de_passe"]):
            st.session_state.authenticated_prof = True
            st.session_state.prof = prof

            enregistrer_log_action(
                email,
                "Connexion Professeur",
                "Connexion réussie"
            )

            return True

    except Exception:
        pass

    return False


def deconnexion_professeur():
    st.session_state.authenticated_prof = False
    st.session_state.prof = None


# ---------- PARENT ----------

def connexion_parent(telephone, annee_naissance):
    try:
        res = (
            supabase.table("parents")
            .select("*")
            .eq("telephone", telephone)
            .eq("annee_naissance", annee_naissance)
            .limit(1)
            .execute()
        )

        if not res.data:
            return False

        parent = res.data[0]

        st.session_state.authenticated_parent = True
        st.session_state.parent = parent

        enregistrer_log_action(
            telephone,
            "Connexion Parent",
            "Connexion réussie"
        )

        return True

    except Exception:
        pass

    return False


def deconnexion_parent():
    st.session_state.authenticated_parent = False
    st.session_state.parent = None


# ---------- INSCRIPTION PROFESSEUR ----------

def creer_professeur(
    nom,
    prenom,
    email,
    matiere,
    classe,
    mot_de_passe
):
    data = {
        "nom": nom,
        "prenom": prenom,
        "email": email,
        "matiere_principale": matiere,
        "classe_attribuee": classe,
        "mot_de_passe": hacher_mot_de_passe(mot_de_passe)
    }

    return sauvegarder_ligne(TABLE_PROFESSEURS, data)


# ---------- INSCRIPTION PARENT ----------

def creer_parent(
    telephone,
    prenom_eleve,
    nom_eleve,
    annee_naissance,
    classe
):
    data = {
        "telephone": telephone,
        "prenom_eleve": prenom_eleve,
        "nom_eleve": nom_eleve,
        "annee_naissance": annee_naissance,
        "classe": classe
    }

    return sauvegarder_ligne("parents", data)


# ---------- VÉRIFICATION SESSION ----------

def admin_connecte():
    return st.session_state.get("authenticated_admin", False)


def professeur_connecte():
    return st.session_state.get("authenticated_prof", False)


def parent_connecte():
    return st.session_state.get("authenticated_parent", False)


# ---------- UTILISATEUR ACTUEL ----------

def utilisateur_actuel():
    if admin_connecte():
        return st.session_state.get("admin")

    if professeur_connecte():
        return st.session_state.get("prof")

    if parent_connecte():
        return st.session_state.get("parent")

    return None# ==========================================
# 6. GESTION DES CLASSES, MATIÈRES ET VIE SCOLAIRE (SUPABASE)
# ==========================================

# ---------- CLASSES ----------

def charger_classes():
    return charger_table(TABLE_CLASSES)


def ajouter_classe(classe, cycle, professeur_responsable):
    data = {
        "classe": classe,
        "cycle": cycle,
        "professeur_responsable": professeur_responsable,
    }
    return sauvegarder_ligne(TABLE_CLASSES, data)


def modifier_classe(id_classe, data):
    return mettre_a_jour_ligne(TABLE_CLASSES, {"id": id_classe}, data)


def supprimer_classe(id_classe):
    return supprimer_ligne(TABLE_CLASSES, {"id": id_classe})


# ---------- MATIÈRES ----------

def charger_matieres(classe=None):
    try:
        query = supabase.table(TABLE_MATIERES).select("*")
        if classe:
            query = query.eq("classe", classe)
        res = query.execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def ajouter_matiere(classe, matiere, coefficient, bareme):
    data = {
        "classe": classe,
        "matiere": matiere,
        "coefficient": coefficient,
        "bareme": bareme,
    }
    return sauvegarder_ligne(TABLE_MATIERES, data)


def modifier_matiere(id_matiere, data):
    return mettre_a_jour_ligne(TABLE_MATIERES, {"id": id_matiere}, data)


def supprimer_matiere(id_matiere):
    return supprimer_ligne(TABLE_MATIERES, {"id": id_matiere})


# ---------- PÉRIODES ----------

def charger_periodes(cycle=None):
    try:
        query = supabase.table(TABLE_PERIODES).select("*")
        if cycle:
            query = query.eq("cycle", cycle)
        res = query.execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def ajouter_periode(periode, cycle, statut="Ouvert"):
    data = {
        "periode": periode,
        "cycle": cycle,
        "statut": statut,
    }
    return sauvegarder_ligne(TABLE_PERIODES, data)


def modifier_periode(id_periode, data):
    return mettre_a_jour_ligne(TABLE_PERIODES, {"id": id_periode}, data)


def supprimer_periode(id_periode):
    return supprimer_ligne(TABLE_PERIODES, {"id": id_periode})


# ---------- VIE SCOLAIRE ----------

def charger_vie_scolaire(classe=None, periode=None):
    try:
        query = supabase.table(TABLE_VIE_SCOLAIRE).select("*")
        if classe:
            query = query.eq("classe", classe)
        if periode:
            query = query.eq("periode", periode)
        res = query.execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def enregistrer_vie_scolaire(
    classe,
    periode,
    eleve,
    absences_justifiees,
    absences_non_justifiees,
    retards,
    heures_perdues,
    observations,
    decision_conseil,
):
    try:
        exist = (
            supabase.table(TABLE_VIE_SCOLAIRE)
            .select("id")
            .eq("classe", classe)
            .eq("periode", periode)
            .eq("eleve", eleve)
            .limit(1)
            .execute()
        )

        data = {
            "classe": classe,
            "periode": periode,
            "eleve": eleve,
            "absences_justifiees": absences_justifiees,
            "absences_non_justifiees": absences_non_justifiees,
            "retards": retards,
            "heures_perdues": heures_perdues,
            "observations": observations,
            "decision_conseil": decision_conseil,
        }

        if exist.data:
            supabase.table(TABLE_VIE_SCOLAIRE).update(data).eq(
                "id", exist.data[0]["id"]
            ).execute()
        else:
            supabase.table(TABLE_VIE_SCOLAIRE).insert(data).execute()

        return True

    except Exception:
        return False


def supprimer_vie_scolaire(id_vie):
    return supprimer_ligne(TABLE_VIE_SCOLAIRE, {"id": id_vie})


# ---------- STATISTIQUES ----------

def statistiques_classe(classe):
    eleves = charger_eleves(classe)
    notes = charger_notes(classe)

    return {
        "effectif": len(eleves),
        "nombre_notes": len(notes),
        "moyenne_generale": round(
            notes["composition"].astype(float).mean(), 2
        ) if not notes.empty else 0,
    }# ==========================================
# 7. TRAVAUX, CAHIER DE TEXTES, MESSAGERIE ET EDT (SUPABASE)
# ==========================================

# ---------- TRAVAUX À FAIRE ----------

def charger_travaux(classe=None, professeur=None):
    try:
        query = supabase.table(TABLE_TRAVAILS).select("*")
        if classe:
            query = query.eq("classe", classe)
        if professeur:
            query = query.eq("professeur", professeur)
        res = query.order("date_publication", desc=True).execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def publier_travail(professeur, date_publication, date_rendu,
                    classe, matiere, titre, consignes,
                    lien_url="", lien_video="",
                    fichier_nom="", fichier_b64="", fichier_type=""):
    data = {
        "professeur": professeur,
        "date_publication": date_publication,
        "date_rendu": date_rendu,
        "classe": classe,
        "matiere": matiere,
        "titre": titre,
        "consignes": consignes,
        "lien_url": lien_url,
        "lien_video": lien_video,
        "fichier_nom": fichier_nom,
        "fichier_b64": fichier_b64,
        "fichier_type": fichier_type,
    }
    return sauvegarder_ligne(TABLE_TRAVAILS, data)


def modifier_travail(id_travail, data):
    return mettre_a_jour_ligne(TABLE_TRAVAILS, {"id": id_travail}, data)


def supprimer_travail(id_travail):
    return supprimer_ligne(TABLE_TRAVAILS, {"id": id_travail})


# ---------- CAHIER DE TEXTES ----------

TABLE_CAHIER_TEXTES = "cahier_textes"


def charger_cahier_textes(classe=None, professeur=None):
    try:
        query = supabase.table(TABLE_CAHIER_TEXTES).select("*")
        if classe:
            query = query.eq("classe", classe)
        if professeur:
            query = query.eq("professeur", professeur)
        res = query.order("date", desc=True).execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def enregistrer_cahier_texte(professeur, date, classe,
                             matiere, contenu, travail_a_faire):
    data = {
        "professeur": professeur,
        "date": date,
        "classe": classe,
        "matiere": matiere,
        "contenu": contenu,
        "travail_a_faire": travail_a_faire,
    }
    return sauvegarder_ligne(TABLE_CAHIER_TEXTES, data)


def modifier_cahier_texte(id_cahier, data):
    return mettre_a_jour_ligne(TABLE_CAHIER_TEXTES, {"id": id_cahier}, data)


def supprimer_cahier_texte(id_cahier):
    return supprimer_ligne(TABLE_CAHIER_TEXTES, {"id": id_cahier})


# ---------- MESSAGERIE PARENTS ----------

def charger_messages(classe=None):
    try:
        query = supabase.table(TABLE_MESSAGES).select("*")
        if classe:
            query = query.eq("classe", classe)
        res = query.order("date_envoi", desc=True).execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def envoyer_message(emetteur, role_emetteur, classe,
                    objet, message, urgent=False):
    data = {
        "emetteur": emetteur,
        "role_emetteur": role_emetteur,
        "date_envoi": datetime.now().isoformat(),
        "classe": classe,
        "objet": objet,
        "message": message,
        "urgent": urgent,
    }
    return sauvegarder_ligne(TABLE_MESSAGES, data)


def supprimer_message(id_message):
    return supprimer_ligne(TABLE_MESSAGES, {"id": id_message})


# ---------- EMPLOI DU TEMPS ----------

TABLE_EDT = "emploi_du_temps"

JOURS_LIST = [
    "Lundi", "Mardi", "Mercredi",
    "Jeudi", "Vendredi", "Samedi"
]

HEURES_LIST = [
    "08h-09h", "09h-10h", "10h-11h",
    "11h00-11h30", "11h30-12h",
    "12h-13h", "13h-14h", "14h-15h",
    "15h-16h", "16h-17h", "17h-18h",
    "18h-19h"
]


def charger_edt(classe):
    try:
        res = (
            supabase.table(TABLE_EDT)
            .select("*")
            .eq("classe", classe)
            .execute()
        )

        df = pd.DataFrame("", index=JOURS_LIST, columns=HEURES_LIST)

        if "11h00-11h30" in df.columns:
            df["11h00-11h30"] = "Récréation"

        if res.data:
            for ligne in res.data:
                jour = ligne["jour"]
                heure = ligne["heure"]
                valeur = (
                    f"{ligne.get('matiere','')} - "
                    f"{ligne.get('professeur','')}"
                )

                if jour in df.index and heure in df.columns:
                    df.loc[jour, heure] = valeur

        return df

    except Exception:
        df = pd.DataFrame("", index=JOURS_LIST, columns=HEURES_LIST)
        if "11h00-11h30" in df.columns:
            df["11h00-11h30"] = "Récréation"
        return df


def enregistrer_edt(classe, jour, heure,
                    matiere, professeur):
    try:
        exist = (
            supabase.table(TABLE_EDT)
            .select("id")
            .eq("classe", classe)
            .eq("jour", jour)
            .eq("heure", heure)
            .limit(1)
            .execute()
        )

        data = {
            "classe": classe,
            "jour": jour,
            "heure": heure,
            "matiere": matiere,
            "professeur": professeur,
        }

        if exist.data:
            supabase.table(TABLE_EDT).update(data).eq(
                "id", exist.data[0]["id"]
            ).execute()
        else:
            supabase.table(TABLE_EDT).insert(data).execute()

        return True

    except Exception:
        return False


def supprimer_creneau_edt(classe, jour, heure):
    try:
        supabase.table(TABLE_EDT).delete() \
            .eq("classe", classe) \
            .eq("jour", jour) \
            .eq("heure", heure) \
            .execute()
        return True
    except Exception:
        return False# ==========================================
# 8. TABLEAU DE BORD & SYNCHRONISATION (SUPABASE)
# ==========================================

# ---------- TABLEAU DE BORD ----------

def tableau_bord():
    try:
        nb_eleves = supabase.table(TABLE_ELEVES).select("id", count="exact").execute().count
        nb_classes = supabase.table(TABLE_CLASSES).select("id", count="exact").execute().count
        nb_profs = supabase.table(TABLE_PROFESSEURS).select("id", count="exact").execute().count
        nb_notes = supabase.table(TABLE_NOTES).select("id", count="exact").execute().count

        moyenne = 0.0
        notes = supabase.table(TABLE_NOTES).select("composition").execute().data
        if notes:
            valeurs = [float(n["composition"]) for n in notes if n.get("composition") is not None]
            if valeurs:
                moyenne = round(sum(valeurs) / len(valeurs), 2)

        return {
            "eleves": nb_eleves or 0,
            "classes": nb_classes or 0,
            "professeurs": nb_profs or 0,
            "notes": nb_notes or 0,
            "moyenne_generale": moyenne,
        }

    except Exception:
        return {
            "eleves": 0,
            "classes": 0,
            "professeurs": 0,
            "notes": 0,
            "moyenne_generale": 0.0,
        }


# ---------- STATISTIQUES PAR CLASSE ----------

def statistiques_par_classe(classe):
    try:
        eleves = charger_eleves(classe)
        notes = charger_notes(classe)

        moyennes = []

        if not eleves.empty:
            periodes = obtenir_periodes_pour_classe(classe)
            periode = periodes[0] if periodes else None

            for _, e in eleves.iterrows():
                m = calculer_moyenne_eleve(classe, e["nom_complet"], periode)
                moyennes.append(m)

        return {
            "effectif": len(eleves),
            "moyenne": round(sum(moyennes) / len(moyennes), 2) if moyennes else 0,
            "meilleure": max(moyennes) if moyennes else 0,
            "plus_faible": min(moyennes) if moyennes else 0,
            "notes": len(notes),
        }

    except Exception:
        return {
            "effectif": 0,
            "moyenne": 0,
            "meilleure": 0,
            "plus_faible": 0,
            "notes": 0,
        }


# ---------- SYNCHRONISATION ----------

def synchroniser_depuis_supabase():
    st.session_state.admin_credentials = charger_table(TABLE_ADMINS)
    st.session_state.prof_credentials = charger_table(TABLE_PROFESSEURS)
    st.session_state.eleves_db = charger_table(TABLE_ELEVES)
    st.session_state.classes_db = charger_table(TABLE_CLASSES)
    st.session_state.matieres_def = charger_table(TABLE_MATIERES)
    st.session_state.notes_db = charger_table(TABLE_NOTES)
    st.session_state.periodes_db = charger_table(TABLE_PERIODES)
    st.session_state.viescolaire_db = charger_table(TABLE_VIE_SCOLAIRE)
    st.session_state.travail_a_faire_db = charger_table(TABLE_TRAVAILS)
    st.session_state.messages_parents_db = charger_table(TABLE_MESSAGES)
    st.session_state.audit_logs_db = charger_table(TABLE_AUDIT_LOGS)


def actualiser_donnees():
    synchroniser_depuis_supabase()
    st.success("Données synchronisées avec Supabase.")


# ---------- SAUVEGARDE AUTOMATIQUE ----------

def sauvegarde_automatique():
    try:
        synchroniser_depuis_supabase()
        return True
    except Exception:
        return False


# ---------- INDICATEURS ----------

def indicateurs_globaux():
    stats = tableau_bord()

    col1, col2, col3, col4, col5 = st.columns(5)

    col1.metric("Élèves", stats["eleves"])
    col2.metric("Classes", stats["classes"])
    col3.metric("Professeurs", stats["professeurs"])
    col4.metric("Notes", stats["notes"])
    col5.metric("Moyenne", stats["moyenne_generale"])


# ---------- LOGS ----------

def afficher_logs():
    logs = charger_table(TABLE_AUDIT_LOGS)

    if logs.empty:
        st.info("Aucun journal disponible.")
        return

    st.dataframe(
        logs.sort_values(by="horodatage", ascending=False),
        use_container_width=True,
    )# ==========================================
# 9. INTERFACE STREAMLIT (SUPABASE)
# ==========================================

MENU = [
    "🏠 Accueil",
    "👨‍💼 Administration",
    "👨‍🏫 Professeurs",
    "👨‍👩‍👧 Parents",
    "🎓 Élèves",
    "📝 Notes",
    "📄 Bulletins",
    "📊 Tableau de bord",
]

st.sidebar.title("École Président Nelson Mandela")
page = st.sidebar.radio("Navigation", MENU)

# ---------- ACCUEIL ----------

if page == "🏠 Accueil":
    st.title("Portail Éducatif National")
    st.subheader("École Président Nelson Mandela")
    stats = tableau_bord()
    col1, col2, col3 = st.columns(3)
    col1.metric("Élèves", stats["eleves"])
    col2.metric("Classes", stats["classes"])
    col3.metric("Professeurs", stats["professeurs"])

# ---------- ADMINISTRATION ----------

elif page == "👨‍💼 Administration":
    st.header("Connexion Administration")

    email = st.text_input("Email")
    mdp = st.text_input("Mot de passe", type="password")

    if st.button("Se connecter"):
        if connexion_admin(email, mdp):
            st.success("Connexion réussie")
        else:
            st.error("Identifiants incorrects")

    if admin_connecte():
        st.subheader("Gestion des classes")
        classes = charger_classes()
        st.dataframe(classes, use_container_width=True)

# ---------- PROFESSEURS ----------

elif page == "👨‍🏫 Professeurs":
    st.header("Espace Professeur")

    email = st.text_input("Email professeur")
    mdp = st.text_input("Mot de passe professeur", type="password")

    if st.button("Connexion professeur"):
        if connexion_professeur(email, mdp):
            st.success("Connexion réussie")
        else:
            st.error("Erreur de connexion")

    if professeur_connecte():
        prof = utilisateur_actuel()
        st.write(f"Bienvenue {prof['prenom']} {prof['nom']}")

        st.subheader("Publier un travail")

        titre = st.text_input("Titre")
        consignes = st.text_area("Consignes")
        classe = st.text_input("Classe")
        matiere = st.text_input("Matière")

        if st.button("Publier"):
            publier_travail(
                professeur=f"{prof['prenom']} {prof['nom']}",
                date_publication=datetime.now().date().isoformat(),
                date_rendu=datetime.now().date().isoformat(),
                classe=classe,
                matiere=matiere,
                titre=titre,
                consignes=consignes,
            )
            st.success("Travail publié")

# ---------- PARENTS ----------

elif page == "👨‍👩‍👧 Parents":
    st.header("Espace Parent")

    tel = st.text_input("Téléphone")
    annee = st.text_input("Année de naissance")

    if st.button("Connexion parent"):
        if connexion_parent(tel, annee):
            st.success("Connexion réussie")
        else:
            st.error("Erreur")

    if parent_connecte():
        parent = utilisateur_actuel()
        st.write(
            f"Élève : {parent['prenom_eleve']} {parent['nom_eleve']}"
        )

        travaux = charger_travaux(parent["classe"])
        st.dataframe(travaux, use_container_width=True)

# ---------- ÉLÈVES ----------

elif page == "🎓 Élèves":
    st.header("Gestion des élèves")

    classes = charger_classes()
    liste_classes = (
        classes["classe"].tolist() if not classes.empty else []
    )

    classe = st.selectbox("Classe", liste_classes)

    eleves = charger_eleves(classe)
    st.dataframe(eleves, use_container_width=True)

# ---------- NOTES ----------

elif page == "📝 Notes":
    st.header("Saisie des notes")

    classes = charger_classes()
    liste_classes = (
        classes["classe"].tolist() if not classes.empty else []
    )

    classe = st.selectbox("Classe", liste_classes)
    periode = st.selectbox(
        "Période", obtenir_periodes_pour_classe(classe)
    )

    eleves = charger_eleves(classe)
    liste_eleves = (
        eleves["nom_complet"].tolist() if not eleves.empty else []
    )

    eleve = st.selectbox("Élève", liste_eleves)

    matieres = charger_matieres(classe)
    liste_matieres = (
        matieres["matiere"].tolist() if not matieres.empty else []
    )

    matiere = st.selectbox("Matière", liste_matieres)

    d1 = st.number_input("Devoir 1", 0.0, 20.0)
    d2 = st.number_input("Devoir 2", 0.0, 20.0)
    comp = st.number_input("Composition", 0.0, 20.0)

    if st.button("Enregistrer la note"):
        enregistrer_note(
            classe,
            periode,
            eleve,
            matiere,
            d1,
            d2,
            comp,
        )
        st.success("Note enregistrée")

# ---------- BULLETINS ----------

elif page == "📄 Bulletins":
    st.header("Génération des bulletins")

    classes = charger_classes()
    liste_classes = (
        classes["classe"].tolist() if not classes.empty else []
    )

    classe = st.selectbox("Classe", liste_classes)
    periode = st.selectbox(
        "Période", obtenir_periodes_pour_classe(classe)
    )

    eleves = charger_eleves(classe)
    liste_eleves = (
        eleves["nom_complet"].tolist() if not eleves.empty else []
    )

    eleve = st.selectbox("Élève", liste_eleves)

    if st.button("Générer le bulletin"):
        bulletin = calculer_bulletin_eleve(
            classe,
            eleve,
            periode,
        )

        if bulletin:
            pdf = generer_pdf_bulletin(bulletin)

            st.download_button(
                "Télécharger le PDF",
                pdf,
                file_name=f"bulletin_{eleve}.pdf",
                mime="application/pdf",
            )
        else:
            st.error("Aucune note disponible")

# ---------- TABLEAU DE BORD ----------

elif page == "📊 Tableau de bord":
    st.header("Tableau de bord")

    indicateurs_globaux()

    if st.button("Synchroniser avec Supabase"):
        actualiser_donnees()

    st.subheader("Journaux d'activité")
    afficher_logs()# ==========================================
# 10. INITIALISATION FINALE & DÉMARRAGE (SUPABASE)
# ==========================================

# ---------- Vérification de la connexion ----------

def verifier_connexion_supabase():
    try:
        supabase.table(TABLE_ADMINS).select("id").limit(1).execute()
        return True
    except Exception as e:
        st.error(f"Connexion Supabase impossible : {e}")
        return False


# ---------- Création des données minimales ----------

def initialiser_donnees_minimales():
    try:
        # Administrateur
        admin = (
            supabase.table(TABLE_ADMINS)
            .select("id")
            .eq("email", ADMIN_EMAIL)
            .limit(1)
            .execute()
        )

        if not admin.data:
            supabase.table(TABLE_ADMINS).insert({
                "nom": "Principal",
                "prenom": "Admin",
                "email": ADMIN_EMAIL,
                "mot_de_passe": hacher_mot_de_passe("cpnm2026"),
                "niveau_acces": "Super-Admin Ayant-Droit",
            }).execute()

        # Classe
        classe = (
            supabase.table(TABLE_CLASSES)
            .select("id")
            .eq("classe", "6ème A")
            .limit(1)
            .execute()
        )

        if not classe.data:
            supabase.table(TABLE_CLASSES).insert({
                "classe": "6ème A",
                "cycle": "Collège",
                "professeur_responsable": "Prof. Math",
            }).execute()

        # Périodes
        periodes = [
            {"periode": "1er Semestre", "cycle": "Collège", "statut": "Ouvert"},
            {"periode": "2ème Semestre", "cycle": "Collège", "statut": "Fermé"},
            {"periode": "1er Trimestre", "cycle": "Élémentaire", "statut": "Ouvert"},
            {"periode": "2ème Trimestre", "cycle": "Élémentaire", "statut": "Fermé"},
            {"periode": "3ème Trimestre", "cycle": "Élémentaire", "statut": "Fermé"},
        ]

        for p in periodes:
            exist = (
                supabase.table(TABLE_PERIODES)
                .select("id")
                .eq("periode", p["periode"])
                .eq("cycle", p["cycle"])
                .limit(1)
                .execute()
            )

            if not exist.data:
                supabase.table(TABLE_PERIODES).insert(p).execute()

    except Exception:
        pass


# ---------- Initialisation des sessions ----------

def initialiser_sessions():
    defaults = {
        "authenticated_admin": False,
        "authenticated_prof": False,
        "authenticated_parent": False,
        "admin": None,
        "prof": None,
        "parent": None,
        "espace_actif": "🏠 Accueil",
        "edt_documents": {},
    }

    for cle, valeur in defaults.items():
        if cle not in st.session_state:
            st.session_state[cle] = valeur


# ---------- Synchronisation ----------

def charger_donnees_application():
    synchroniser_depuis_supabase()


# ---------- Démarrage ----------

def demarrer_application():
    initialiser_sessions()

    if not verifier_connexion_supabase():
        st.stop()

    initialiser_donnees_minimales()
    charger_donnees_application()


# ---------- Exécution ----------

demarrer_application()-- ==========================================
-- 11. CRÉATION DES TABLES SUPABASE
-- ==========================================

create extension if not exists pgcrypto;

-- ADMINS
create table if not exists admins (
    id uuid primary key default gen_random_uuid(),
    nom text,
    prenom text,
    email text unique,
    mot_de_passe text,
    niveau_acces text,
    created_at timestamp default now()
);

-- PROFESSEURS
create table if not exists professeurs (
    id uuid primary key default gen_random_uuid(),
    nom text,
    prenom text,
    email text unique,
    matiere_principale text,
    classe_attribuee text,
    mot_de_passe text,
    created_at timestamp default now()
);

-- PARENTS
create table if not exists parents (
    id uuid primary key default gen_random_uuid(),
    telephone text,
    prenom_eleve text,
    nom_eleve text,
    annee_naissance text,
    classe text,
    created_at timestamp default now()
);

-- CLASSES
create table if not exists classes (
    id uuid primary key default gen_random_uuid(),
    classe text unique,
    cycle text,
    professeur_responsable text,
    created_at timestamp default now()
);

-- ÉLÈVES
create table if not exists eleves (
    id uuid primary key default gen_random_uuid(),
    nom_complet text,
    prenom text,
    nom text,
    date_naissance date,
    classe text,
    photo text,
    created_at timestamp default now()
);

-- MATIÈRES
create table if not exists matieres (
    id uuid primary key default gen_random_uuid(),
    classe text,
    matiere text,
    coefficient numeric default 1,
    bareme numeric default 20,
    created_at timestamp default now()
);

-- PÉRIODES
create table if not exists periodes (
    id uuid primary key default gen_random_uuid(),
    periode text,
    cycle text,
    statut text default 'Ouvert',
    created_at timestamp default now()
);

-- NOTES
create table if not exists notes (
    id uuid primary key default gen_random_uuid(),
    classe text,
    periode text,
    eleve text,
    matiere text,
    devoir1 numeric default 0,
    devoir2 numeric default 0,
    composition numeric default 0,
    created_at timestamp default now()
);

-- VIE SCOLAIRE
create table if not exists vie_scolaire (
    id uuid primary key default gen_random_uuid(),
    classe text,
    periode text,
    eleve text,
    absences_justifiees integer default 0,
    absences_non_justifiees integer default 0,
    retards integer default 0,
    heures_perdues integer default 0,
    observations text,
    decision_conseil text,
    created_at timestamp default now()
);

-- TRAVAUX
create table if not exists travaux (
    id uuid primary key default gen_random_uuid(),
    professeur text,
    date_publication date,
    date_rendu date,
    classe text,
    matiere text,
    titre text,
    consignes text,
    lien_url text,
    lien_video text,
    fichier_nom text,
    fichier_b64 text,
    fichier_type text,
    created_at timestamp default now()
);

-- CAHIER DE TEXTES
create table if not exists cahier_textes (
    id uuid primary key default gen_random_uuid(),
    professeur text,
    date date,
    classe text,
    matiere text,
    contenu text,
    travail_a_faire text,
    created_at timestamp default now()
);

-- MESSAGES
create table if not exists messages (
    id uuid primary key default gen_random_uuid(),
    emetteur text,
    role_emetteur text,
    date_envoi timestamp,
    classe text,
    objet text,
    message text,
    urgent boolean default false,
    created_at timestamp default now()
);

-- EMPLOI DU TEMPS
create table if not exists emploi_du_temps (
    id uuid primary key default gen_random_uuid(),
    classe text,
    jour text,
    heure text,
    matiere text,
    professeur text,
    created_at timestamp default now()
);

-- AUDIT LOGS
create table if not exists audit_logs (
    id uuid primary key default gen_random_uuid(),
    horodatage timestamp,
    acteur text,
    action text,
    details text,
    created_at timestamp default now()
);

-- INDEX
create index if not exists idx_notes_classe on notes(classe);
create index if not exists idx_notes_eleve on notes(eleve);
create index if not exists idx_notes_periode on notes(periode);
create index if not exists idx_eleves_classe on eleves(classe);
create index if not exists idx_vie_scolaire on vie_scolaire(classe, periode, eleve);-- ==========================================
-- 12. SÉCURITÉ SUPABASE (RLS)
-- ==========================================

-- Activer RLS sur toutes les tables
alter table admins enable row level security;
alter table professeurs enable row level security;
alter table parents enable row level security;
alter table classes enable row level security;
alter table eleves enable row level security;
alter table matieres enable row level security;
alter table periodes enable row level security;
alter table notes enable row level security;
alter table vie_scolaire enable row level security;
alter table travaux enable row level security;
alter table cahier_textes enable row level security;
alter table messages enable row level security;
alter table emploi_du_temps enable row level security;
alter table audit_logs enable row level security;

-- Lecture publique (clé anon)
create policy "lecture_classes"
on classes for select
using (true);

create policy "lecture_matieres"
on matieres for select
using (true);

create policy "lecture_periodes"
on periodes for select
using (true);

create policy "lecture_eleves"
on eleves for select
using (true);

create policy "lecture_notes"
on notes for select
using (true);

create policy "lecture_vie_scolaire"
on vie_scolaire for select
using (true);

create policy "lecture_travaux"
on travaux for select
using (true);

create policy "lecture_cahier_textes"
on cahier_textes for select
using (true);

create policy "lecture_messages"
on messages for select
using (true);

create policy "lecture_edt"
on emploi_du_temps for select
using (true);

create policy "lecture_admins"
on admins for select
using (true);

create policy "lecture_professeurs"
on professeurs for select
using (true);

create policy "lecture_parents"
on parents for select
using (true);

create policy "lecture_audit_logs"
on audit_logs for select
using (true);

-- Écriture (insert)
create policy "insert_all"
on classes for insert with check (true);

create policy "insert_matieres"
on matieres for insert with check (true);

create policy "insert_eleves"
on eleves for insert with check (true);

create policy "insert_notes"
on notes for insert with check (true);

create policy "insert_vie_scolaire"
on vie_scolaire for insert with check (true);

create policy "insert_travaux"
on travaux for insert with check (true);

create policy "insert_cahier"
on cahier_textes for insert with check (true);

create policy "insert_messages"
on messages for insert with check (true);

create policy "insert_edt"
on emploi_du_temps for insert with check (true);

create policy "insert_admins"
on admins for insert with check (true);

create policy "insert_professeurs"
on professeurs for insert with check (true);

create policy "insert_parents"
on parents for insert with check (true);

create policy "insert_logs"
on audit_logs for insert with check (true);

-- Modification
create policy "update_all"
on classes for update
using (true);

create policy "update_matieres"
on matieres for update
using (true);

create policy "update_eleves"
on eleves for update
using (true);

create policy "update_notes"
on notes for update
using (true);

create policy "update_vie_scolaire"
on vie_scolaire for update
using (true);

create policy "update_travaux"
on travaux for update
using (true);

create policy "update_cahier"
on cahier_textes for update
using (true);

create policy "update_messages"
on messages for update
using (true);

create policy "update_edt"
on emploi_du_temps for update
using (true);

create policy "update_admins"
on admins for update
using (true);

create policy "update_professeurs"
on professeurs for update
using (true);

create policy "update_parents"
on parents for update
using (true);

create policy "update_logs"
on audit_logs for update
using (true);

-- Suppression
create policy "delete_all"
on classes for delete
using (true);

create policy "delete_matieres"
on matieres for delete
using (true);

create policy "delete_eleves"
on eleves for delete
using (true);

create policy "delete_notes"
on notes for delete
using (true);

create policy "delete_vie_scolaire"
on vie_scolaire for delete
using (true);

create policy "delete_travaux"
on travaux for delete
using (true);

create policy "delete_cahier"
on cahier_textes for delete
using (true);

create policy "delete_messages"
on messages for delete
using (true);

create policy "delete_edt"
on emploi_du_temps for delete
using (true);

create policy "delete_admins"
on admins for delete
using (true);

create policy "delete_professeurs"
on professeurs for delete
using (true);

create policy "delete_parents"
on parents for delete
using (true);

create policy "delete_logs"
on audit_logs for delete
using (true);# ==========================================
# 13. CONFIGURATION SÉCURISÉE (STREAMLIT CLOUD)
# ==========================================

import streamlit as st
from supabase import create_client, Client

# Variables d'environnement Streamlit
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Test de connexion
def test_connexion():
    try:
        supabase.table("classes").select("id").limit(1).execute()
        return True
    except Exception as e:
        st.error(f"Erreur Supabase : {e}")
        return False

if not test_connexion():
    st.stop()# ==========================================
# 13. MIGRATION GLOBALE VERS SUPABASE
# ==========================================

# ---------- Actualisation globale ----------

def actualiser():
    st.session_state.admin_credentials = charger_table(TABLE_ADMINS)
    st.session_state.prof_credentials = charger_table(TABLE_PROFESSEURS)
    st.session_state.eleves_db = charger_table(TABLE_ELEVES)
    st.session_state.classes_db = charger_table(TABLE_CLASSES)
    st.session_state.matieres_def = charger_table(TABLE_MATIERES)
    st.session_state.notes_db = charger_table(TABLE_NOTES)
    st.session_state.periodes_db = charger_table(TABLE_PERIODES)
    st.session_state.viescolaire_db = charger_table(TABLE_VIE_SCOLAIRE)
    st.session_state.travail_a_faire_db = charger_table(TABLE_TRAVAILS)
    st.session_state.messages_parents_db = charger_table(TABLE_MESSAGES)
    st.session_state.audit_logs_db = charger_table(TABLE_AUDIT_LOGS)


# ---------- CRUD universel ----------

def inserer(table, data):
    try:
        supabase.table(table).insert(data).execute()
        actualiser()
        return True
    except Exception as e:
        st.error(e)
        return False


def modifier(table, id_ligne, data):
    try:
        supabase.table(table).update(data).eq("id", id_ligne).execute()
        actualiser()
        return True
    except Exception as e:
        st.error(e)
        return False


def supprimer(table, id_ligne):
    try:
        supabase.table(table).delete().eq("id", id_ligne).execute()
        actualiser()
        return True
    except Exception as e:
        st.error(e)
        return False


# ---------- Import d'élèves ----------

def importer_eleves_dataframe(df):
    if df.empty:
        return False

    try:
        donnees = []

        for _, row in df.iterrows():
            donnees.append({
                "nom_complet": row.get("Nom Complet", ""),
                "prenom": row.get("Prénom", ""),
                "nom": row.get("Nom", ""),
                "date_naissance": row.get("Date de Naissance", None),
                "classe": row.get("Classe", ""),
                "photo": row.get("Photo", "")
            })

        supabase.table(TABLE_ELEVES).insert(donnees).execute()
        actualiser()
        return True

    except Exception as e:
        st.error(e)
        return False


# ---------- Export DataFrame ----------

def exporter_table(table):
    return charger_table(table)


# ---------- Réinitialisation ----------

def recharger_application():
    actualiser()
    st.success("Application synchronisée avec Supabase.")


# ---------- Synchronisation au démarrage ----------

if "supabase_sync" not in st.session_state:
    actualiser()
    st.session_state.supabase_sync = True# ==========================================
# 14. SAUVEGARDE, RESTAURATION & MAINTENANCE
# ==========================================

from io import BytesIO
from openpyxl import Workbook

# ---------- Export Excel ----------

def exporter_excel(table_name):
    df = charger_table(table_name)

    wb = Workbook()
    ws = wb.active
    ws.title = table_name

    if not df.empty:
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


# ---------- Export PDF ----------

def exporter_pdf_table(table_name):
    df = charger_table(table_name)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 8, table_name.upper(), 0, 1, "C")

    if not df.empty:
        pdf.set_font("Arial", "", 8)

        for _, row in df.iterrows():
            ligne = " | ".join(
                [str(v)[:25] for v in row.tolist()]
            )
            pdf.multi_cell(0, 5, ligne)

    return pdf.output(dest="S").encode("latin-1")


# ---------- Sauvegarde JSON ----------

def sauvegarder_json():
    tables = {
        "admins": TABLE_ADMINS,
        "professeurs": TABLE_PROFESSEURS,
        "parents": "parents",
        "classes": TABLE_CLASSES,
        "eleves": TABLE_ELEVES,
        "matieres": TABLE_MATIERES,
        "periodes": TABLE_PERIODES,
        "notes": TABLE_NOTES,
        "vie_scolaire": TABLE_VIE_SCOLAIRE,
        "travaux": TABLE_TRAVAILS,
        "cahier_textes": "cahier_textes",
        "messages": TABLE_MESSAGES,
        "emploi_du_temps": "emploi_du_temps",
        "audit_logs": TABLE_AUDIT_LOGS,
    }

    sauvegarde = {}

    for nom, table in tables.items():
        df = charger_table(table)
        sauvegarde[nom] = (
            df.to_dict(orient="records")
            if not df.empty else []
        )

    return json.dumps(
        sauvegarde,
        ensure_ascii=False,
        indent=2,
        default=str,
    ).encode("utf-8")


# ---------- Restauration JSON ----------

def restaurer_json(data_json):
    try:
        contenu = json.loads(data_json)

        mapping = {
            "admins": TABLE_ADMINS,
            "professeurs": TABLE_PROFESSEURS,
            "parents": "parents",
            "classes": TABLE_CLASSES,
            "eleves": TABLE_ELEVES,
            "matieres": TABLE_MATIERES,
            "periodes": TABLE_PERIODES,
            "notes": TABLE_NOTES,
            "vie_scolaire": TABLE_VIE_SCOLAIRE,
            "travaux": TABLE_TRAVAILS,
            "cahier_textes": "cahier_textes",
            "messages": TABLE_MESSAGES,
            "emploi_du_temps": "emploi_du_temps",
            "audit_logs": TABLE_AUDIT_LOGS,
        }

        for nom, table in mapping.items():
            if nom in contenu:
                donnees = contenu[nom]

                if donnees:
                    supabase.table(table).insert(
                        donnees
                    ).execute()

        actualiser()

        return True

    except Exception as e:
        st.error(e)
        return False


# ---------- Nettoyage ----------

def nettoyer_table(table_name):
    try:
        supabase.table(table_name).delete().neq(
            "id", ""
        ).execute()

        actualiser()

        return True

    except Exception:
        return False


# ---------- Vérification ----------

def verifier_integrite():
    erreurs = []

    tables = [
        TABLE_ADMINS,
        TABLE_PROFESSEURS,
        "parents",
        TABLE_CLASSES,
        TABLE_ELEVES,
        TABLE_MATIERES,
        TABLE_PERIODES,
        TABLE_NOTES,
        TABLE_VIE_SCOLAIRE,
        TABLE_TRAVAILS,
        "cahier_textes",
        TABLE_MESSAGES,
        "emploi_du_temps",
        TABLE_AUDIT_LOGS,
    ]

    for table in tables:
        try:
            supabase.table(table).select(
                "id"
            ).limit(1).execute()
        except Exception:
            erreurs.append(table)

    return erreurs


# ---------- Maintenance ----------

def maintenance():
    erreurs = verifier_integrite()

    if erreurs:
        st.error(
            f"Tables en erreur : {', '.join(erreurs)}"
        )
    else:
        st.success("Base Supabase opérationnelle.")# ==========================================
# 15. INITIALISATION COMPLÈTE DES DONNÉES SUPABASE
# ==========================================

def initialiser_matieres_par_defaut():
    matieres = [
        # Collège
        {"classe": "6ème A", "matiere": "Mathématiques", "coefficient": 4, "bareme": 20},
        {"classe": "6ème A", "matiere": "Français", "coefficient": 5, "bareme": 20},
        {"classe": "6ème A", "matiere": "Histoire-Géographie", "coefficient": 2, "bareme": 20},
        {"classe": "6ème A", "matiere": "SVT", "coefficient": 2, "bareme": 20},
        {"classe": "6ème A", "matiere": "Anglais", "coefficient": 2, "bareme": 20},
        {"classe": "6ème A", "matiere": "Physique-Chimie", "coefficient": 2, "bareme": 20},

        # Élémentaire
        {"classe": "CP", "matiere": "Lecture / Langage", "coefficient": 1, "bareme": 50},
        {"classe": "CP", "matiere": "Calcul / Mathématiques", "coefficient": 1, "bareme": 50},
        {"classe": "CP", "matiere": "Éveil / Science", "coefficient": 1, "bareme": 30},
        {"classe": "CP", "matiere": "Éducation Civique", "coefficient": 1, "bareme": 20},
    ]

    for m in matieres:
        exist = (
            supabase.table(TABLE_MATIERES)
            .select("id")
            .eq("classe", m["classe"])
            .eq("matiere", m["matiere"])
            .limit(1)
            .execute()
        )

        if not exist.data:
            supabase.table(TABLE_MATIERES).insert(m).execute()


def initialiser_classes_par_defaut():
    classes = [
        {"classe": "CP", "cycle": "Élémentaire", "professeur_responsable": "Prof. Élémentaire"},
        {"classe": "6ème A", "cycle": "Collège", "professeur_responsable": "Prof. Math"},
    ]

    for c in classes:
        exist = (
            supabase.table(TABLE_CLASSES)
            .select("id")
            .eq("classe", c["classe"])
            .limit(1)
            .execute()
        )

        if not exist.data:
            supabase.table(TABLE_CLASSES).insert(c).execute()


def initialiser_periodes_par_defaut():
    periodes = [
        {"periode": "1er Trimestre", "cycle": "Élémentaire", "statut": "Ouvert"},
        {"periode": "2ème Trimestre", "cycle": "Élémentaire", "statut": "Fermé"},
        {"periode": "3ème Trimestre", "cycle": "Élémentaire", "statut": "Fermé"},
        {"periode": "1er Semestre", "cycle": "Collège", "statut": "Ouvert"},
        {"periode": "2ème Semestre", "cycle": "Collège", "statut": "Fermé"},
    ]

    for p in periodes:
        exist = (
            supabase.table(TABLE_PERIODES)
            .select("id")
            .eq("periode", p["periode"])
            .eq("cycle", p["cycle"])
            .limit(1)
            .execute()
        )

        if not exist.data:
            supabase.table(TABLE_PERIODES).insert(p).execute()


def creer_eleve_demo():
    exist = (
        supabase.table(TABLE_ELEVES)
        .select("id")
        .eq("nom_complet", "DIOP Moussa")
        .limit(1)
        .execute()
    )

    if not exist.data:
        supabase.table(TABLE_ELEVES).insert({
            "nom_complet": "DIOP Moussa",
            "prenom": "Moussa",
            "nom": "DIOP",
            "date_naissance": "2013-05-10",
            "classe": "6ème A",
            "photo": "",
        }).execute()


def initialisation_complete():
    initialiser_classes_par_defaut()
    initialiser_periodes_par_defaut()
    initialiser_matieres_par_defaut()
    creer_eleve_demo()
    actualiser()
    st.success("Initialisation complète Supabase terminée.")


# Exécution au premier lancement
if "initialisation_supabase" not in st.session_state:
    initialisation_complete()
    st.session_state.initialisation_supabase = True# ==========================================
# 16. GESTION DES UTILISATEURS & PERMISSIONS
# ==========================================

# ---------- ADMINISTRATEURS ----------

def creer_administrateur(nom, prenom, email, mot_de_passe, niveau="Admin"):
    data = {
        "nom": nom,
        "prenom": prenom,
        "email": email,
        "mot_de_passe": hacher_mot_de_passe(mot_de_passe),
        "niveau_acces": niveau,
    }
    return sauvegarder_ligne(TABLE_ADMINS, data)


def liste_administrateurs():
    return charger_table(TABLE_ADMINS)


def supprimer_administrateur(id_admin):
    return supprimer_ligne(TABLE_ADMINS, {"id": id_admin})


# ---------- PROFESSEURS ----------

def creer_professeur_complet(
    nom,
    prenom,
    email,
    matiere,
    classe,
    mot_de_passe,
):
    data = {
        "nom": nom,
        "prenom": prenom,
        "email": email,
        "matiere_principale": matiere,
        "classe_attribuee": classe,
        "mot_de_passe": hacher_mot_de_passe(mot_de_passe),
    }
    return sauvegarder_ligne(TABLE_PROFESSEURS, data)


def liste_professeurs():
    return charger_table(TABLE_PROFESSEURS)


def modifier_professeur(id_prof, data):
    return mettre_a_jour_ligne(TABLE_PROFESSEURS, {"id": id_prof}, data)


def supprimer_professeur(id_prof):
    return supprimer_ligne(TABLE_PROFESSEURS, {"id": id_prof})


# ---------- PARENTS ----------

def creer_parent_complet(
    telephone,
    prenom_eleve,
    nom_eleve,
    annee_naissance,
    classe,
):
    data = {
        "telephone": telephone,
        "prenom_eleve": prenom_eleve,
        "nom_eleve": nom_eleve,
        "annee_naissance": annee_naissance,
        "classe": classe,
    }
    return sauvegarder_ligne("parents", data)


def liste_parents():
    return charger_table("parents")


def supprimer_parent(id_parent):
    return supprimer_ligne("parents", {"id": id_parent})


# ---------- MOT DE PASSE ----------

def reinitialiser_mot_de_passe_admin(email, nouveau_mdp):
    return mettre_a_jour_ligne(
        TABLE_ADMINS,
        {"email": email},
        {"mot_de_passe": hacher_mot_de_passe(nouveau_mdp)},
    )


def reinitialiser_mot_de_passe_prof(email, nouveau_mdp):
    return mettre_a_jour_ligne(
        TABLE_PROFESSEURS,
        {"email": email},
        {"mot_de_passe": hacher_mot_de_passe(nouveau_mdp)},
    )


# ---------- PERMISSIONS ----------

PERMISSIONS = {
    "Super-Admin Ayant-Droit": {
        "gestion_admins": True,
        "gestion_profs": True,
        "gestion_eleves": True,
        "gestion_notes": True,
        "gestion_bulletins": True,
        "gestion_classes": True,
        "gestion_finances": True,
        "gestion_systeme": True,
    },
    "Admin": {
        "gestion_admins": False,
        "gestion_profs": True,
        "gestion_eleves": True,
        "gestion_notes": True,
        "gestion_bulletins": True,
        "gestion_classes": True,
        "gestion_finances": False,
        "gestion_systeme": False,
    },
    "Professeur": {
        "gestion_admins": False,
        "gestion_profs": False,
        "gestion_eleves": False,
        "gestion_notes": True,
        "gestion_bulletins": False,
        "gestion_classes": False,
        "gestion_finances": False,
        "gestion_systeme": False,
    },
    "Parent": {
        "gestion_admins": False,
        "gestion_profs": False,
        "gestion_eleves": False,
        "gestion_notes": False,
        "gestion_bulletins": False,
        "gestion_classes": False,
        "gestion_finances": False,
        "gestion_systeme": False,
    },
}


def utilisateur_a_permission(permission):
    if admin_connecte():
        admin = utilisateur_actuel()
        niveau = admin.get("niveau_acces", "Admin")
        return PERMISSIONS.get(niveau, {}).get(permission, False)

    if professeur_connecte():
        return PERMISSIONS["Professeur"].get(permission, False)

    if parent_connecte():
        return PERMISSIONS["Parent"].get(permission, False)

    return False


# ---------- CONTRÔLE D'ACCÈS ----------

def exiger_permission(permission):
    if not utilisateur_a_permission(permission):
        st.error("Accès refusé.")
        st.stop()


# ---------- JOURNAL ----------

def journal_utilisateur(action, details=""):
    utilisateur = utilisateur_actuel()

    if utilisateur:
        identifiant = (
            utilisateur.get("email")
            or utilisateur.get("telephone")
            or "Utilisateur"
        )
    else:
        identifiant = "Anonyme"

    enregistrer_log_action(identifiant, action, details)
